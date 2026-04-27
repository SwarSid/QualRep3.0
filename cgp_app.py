import streamlit as st
import pandas as pd
import re
from collections import defaultdict
 
st.set_page_config(page_title="CGP Insight Engine", layout="wide", page_icon="🧬")
 
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=Geist:wght@300;400;500;600&display=swap');
html,body,[data-testid="stAppViewContainer"]{background:#080c14!important;color:#c8d4e8;font-family:'Geist',sans-serif}
[data-testid="stHeader"]{background:transparent!important}
.block-container{padding:1.5rem 2rem!important;max-width:1600px!important}
.stTextInput>div>div>input{background:#0d1420!important;border:1px solid #1e2d47!important;border-radius:10px!important;color:#e2ecf8!important;font-size:15px!important;padding:12px 18px!important}
.card{background:#0d1420;border:1px solid #1a2640;border-radius:12px;padding:18px 22px;margin-bottom:14px}
.stat-num{font-family:'IBM Plex Mono',monospace;font-size:38px;font-weight:600;color:#e2ecf8;line-height:1}
.stat-lbl{font-size:10px;color:#4a6080;letter-spacing:1.5px;text-transform:uppercase;margin-top:4px}
.sec-lbl{font-size:10px;letter-spacing:2px;text-transform:uppercase;color:#3b6ef7;font-weight:600;margin-bottom:12px;font-family:'IBM Plex Mono',monospace}
.fq-card{background:#0a1220;border:1px solid #1a2640;border-left:3px solid #3b6ef7;border-radius:0 12px 12px 0;padding:16px 20px;margin-bottom:12px}
.fq-card.acad{border-left-color:#3b6ef7}
.fq-card.comm{border-left-color:#f59e0b}
.fq-card.fmi{border-left-color:#4A4A4A}
.fq-card.tempus{border-left-color:#52C4A0}
.fq-card.caris{border-left-color:#E8745A}
.fq-card.guardant{border-left-color:#9B59B6}
.fq-meta{font-family:'IBM Plex Mono',monospace;font-size:10px;color:#4a6080;margin-bottom:8px}
.fq-text{font-size:13px;color:#c8d4e8;line-height:1.9}
.hl-testing{background:rgba(59,130,246,.25);border-radius:3px;padding:1px 4px;color:#60a5fa;font-style:normal;font-weight:600}
.hl-support{background:rgba(16,185,129,.25);border-radius:3px;padding:1px 4px;color:#34d399;font-style:normal;font-weight:600}
.hl-accuracy{background:rgba(245,158,11,.25);border-radius:3px;padding:1px 4px;color:#fbbf24;font-style:normal;font-weight:600}
.hl-report{background:rgba(139,92,246,.25);border-radius:3px;padding:1px 4px;color:#a78bfa;font-style:normal;font-weight:600}
.hl-cost{background:rgba(249,115,22,.25);border-radius:3px;padding:1px 4px;color:#fb923c;font-style:normal;font-weight:600}
.hl-tat{background:rgba(239,68,68,.25);border-radius:3px;padding:1px 4px;color:#f87171;font-style:normal;font-weight:600}
.hl-emr{background:rgba(132,204,22,.25);border-radius:3px;padding:1px 4px;color:#a3e635;font-style:normal;font-weight:600}
.hl-def{background:rgba(59,110,247,.2);border-radius:3px;padding:1px 4px;color:#93c5fd;font-style:normal;font-weight:600}
.tag{display:inline-block;padding:2px 8px;border-radius:99px;font-size:10px;font-weight:600;margin-right:4px;margin-top:2px}
.bar-row{display:flex;align-items:center;gap:10px;margin-bottom:9px}
.bar-lbl{font-size:11px;color:#8a9ab5;min-width:200px}
.bar-trk{flex:1;height:5px;background:#1a2640;border-radius:3px}
.bar-fill{height:5px;border-radius:3px}
.bar-cnt{font-family:'IBM Plex Mono',monospace;font-size:11px;color:#4a6080;min-width:60px;text-align:right}
.diff-pos{color:#10b981;font-weight:700}
.diff-neg{color:#ef4444;font-weight:700}
.seg-a{background:rgba(245,158,11,.15);color:#fbbf24;border:1px solid rgba(245,158,11,.3);padding:2px 10px;border-radius:99px;font-size:11px;font-weight:600}
.seg-b{background:rgba(59,110,247,.15);color:#93c5fd;border:1px solid rgba(59,110,247,.3);padding:2px 10px;border-radius:99px;font-size:11px;font-weight:600}
.ibadge{display:inline-block;padding:3px 12px;border-radius:99px;font-size:10px;font-weight:600;letter-spacing:1px;text-transform:uppercase;background:rgba(59,110,247,.12);color:#3b6ef7;border:1px solid rgba(59,110,247,.25);margin-right:6px}
.tbadge{display:inline-block;padding:3px 12px;border-radius:99px;font-size:10px;background:rgba(16,185,129,.1);color:#34d399;border:1px solid rgba(16,185,129,.2);margin-right:6px}
[data-testid="stExpander"]{background:#0d1420!important;border:1px solid #1a2640!important;border-radius:10px!important}
.stButton>button{background:#0d1420!important;border:1px solid #1e2d47!important;color:#8a9ab5!important;border-radius:8px!important;font-size:11px!important;padding:5px 10px!important;width:100%!important;text-align:left!important;white-space:normal!important;height:auto!important}
.stButton>button:hover{border-color:#3b6ef7!important;color:#93c5fd!important}
.stSelectbox>div>div{background:#0d1420!important;border:1px solid #1e2d47!important;color:#c8d4e8!important;border-radius:8px!important}
hr{border:none;border-top:1px solid #1a2640;margin:20px 0}
</style>
""", unsafe_allow_html=True)
 
# ─────────────────────────────────────────────────────────────────────────────
# CGP THEMES — mapped from the study topics
# ─────────────────────────────────────────────────────────────────────────────
THEMES = {
"Testing Modality": ["testing modality","test modality","testing type","modality","test type","testing approach","ngs","next generation sequencing","liquid biopsy","tissue testing","ctdna","blood-based","tissue-based","whole genome","sequencing","genomic profiling","cgp","comprehensive genomic"],
"Physician Support Services": ["physician support","medical science liaison","msl","clinical support","rep support","sales rep","representative","support services","medical affairs","physician education","clinical liaison","provider support","account manager"],
"Test Accuracy": ["accuracy","accurate","sensitivity","specificity","false positive","false negative","reliable","reliability","precision","concordance","validated","analytical validity","clinical validity","correct result","error rate"],
"Clinical Trial Matching": ["clinical trial","trial match","trial matching","trial eligibility","trial access","basket trial","umbrella trial","actionable variant","biomarker match","trial enrollment","patient matching","matched therapy"],
"Ease of Ordering": ["ease of ordering","easy to order","ordering process","order submission","simple to order","ordering system","requisition","test ordering","how to order","ordering experience","order online","streamlined ordering"],
"Assay Panel Size": ["panel size","gene panel","number of genes","genes covered","biomarkers covered","comprehensive panel","broad panel","large panel","assay coverage","genes tested","biomarker coverage","panel breadth","350 genes","500 genes","full panel"],
"Clinical Utility": ["clinical utility","clinical relevance","actionable","actionability","treatment decision","clinical impact","changes management","guides treatment","influences decision","meaningful result","useful result","clinical significance","prognostic","predictive","therapeutic implication"],
"Turnaround Time": ["turnaround","tat","turn around","time to result","result time","how long","days to result","reporting time","quick result","fast result","slow","delay","timely","rapid","expedited","within days","two weeks","three weeks"],
"Report Clarity": ["report clarity","report quality","report format","easy to read","readable report","clear report","report layout","report presentation","interpretation","annotated","visual","easy to interpret","user friendly report","report design","result presentation","icd","ihc","result format"],
"Patient Financial Support": ["patient financial","financial support","out of pocket","copay","co-pay","patient assistance","financial assistance","cost to patient","affordability","patient cost","billing","insurance","coverage","reimbursement","financial services","patient support program"],
"Aggregated Patient Data": ["aggregated data","real world data","real-world","database","patient database","data access","population data","cohort","aggregate","benchmark","rwd","patient population data","outcomes data","anonymized data","de-identified"],
"Other Portfolio Tests": ["portfolio","other tests","other products","suite of tests","additional tests","other assays","companion diagnostic","liquid","tissue","multi-test","test menu","full portfolio","complete portfolio","product range"],
"Reflex Testing": ["reflex","reflexing","automatic reflex","reflex test","reflexed","automatic testing","triggered test","cascade testing","sequential testing","add-on test"],
"Research Support": ["research support","research collaboration","publications","peer reviewed","clinical evidence","study data","research partnership","academic collaboration","investigator","research program","scientific support","data sharing"],
"EMR/EHR Integration": ["emr","ehr","electronic medical record","electronic health record","integration","integrated","epic","cerner","ordering through emr","results in emr","portal","app","digital","electronic ordering","seamless","workflow integration"],
}
 
TC = {
"Testing Modality":"#3b82f6","Physician Support Services":"#10b981",
"Test Accuracy":"#f59e0b","Clinical Trial Matching":"#8b5cf6",
"Ease of Ordering":"#84cc16","Assay Panel Size":"#06b6d4",
"Clinical Utility":"#34d399","Turnaround Time":"#ef4444",
"Report Clarity":"#a78bfa","Patient Financial Support":"#f97316",
"Aggregated Patient Data":"#fbbf24","Other Portfolio Tests":"#ec4899",
"Reflex Testing":"#22d3ee","Research Support":"#60a5fa",
"EMR/EHR Integration":"#a3e635",
}
 
HL_MAP = {
"Testing Modality":"hl-testing","Physician Support Services":"hl-support",
"Test Accuracy":"hl-accuracy","Report Clarity":"hl-report",
"Patient Financial Support":"hl-cost","Turnaround Time":"hl-tat",
"EMR/EHR Integration":"hl-emr",
}
 
# Vendor colour mapping for badges
VENDOR_CSS = {
"FMI":"fmi","Tempus":"tempus","Caris Life Sciences":"caris","Guardant Health":"guardant"
}
VENDOR_COLORS = {
"FMI":"#4A4A4A",
"Guardant Health":"#9B59B6",
"Caris Life Sciences":"#E8745A",
"Tempus":"#52C4A0",
}
 
# ─────────────────────────────────────────────────────────────────────────────
# DATA LOADER
# ─────────────────────────────────────────────────────────────────────────────
INVALID_PATTERNS = [
r'^(yes|no|ok|okay|sure|good|great|fine|same|n/a|na|none|nothing|not sure|idk|unclear)\.?$',
r'^(six|seven|eight|nine|ten|six,? ?seven)\.?$',
r'^(goo|goood|gret|grea)\.?$',
r'^\d+\.?$',
]
INVALID_RE = re.compile('|'.join(INVALID_PATTERNS), re.IGNORECASE)
 
def simplify_setting(v):
vu = str(v).upper()
if 'ACADEMIC' in vu or 'MAJOR' in vu: return 'Academic'
if 'COMMUNITY' in vu: return 'Community'
if 'OFFICE' in vu or 'PRIVATE' in vu: return 'Private Practice'
if 'TEACHING' in vu and 'AFFILIATED' in vu: return 'Teaching Hospital'
if 'VA' in vu or 'MILITARY' in vu: return 'VA/Military'
return v
 
def load_excel(file):
try:
import openpyxl
wb = openpyxl.load_workbook(file, read_only=True)
ws = wb.active
all_rows = list(ws.iter_rows(values_only=True))
wb.close()
except Exception as e:
return None, f"Cannot open: {e}"
 
if not all_rows:
return None, "File is empty."
 
headers = all_rows[0]
data_rows = all_rows[1:]
 
# Find text column — longest avg
import pandas as pd
raw = pd.DataFrame(data_rows, columns=[str(h) if h else f"col_{i}" for i, h in enumerate(headers)])
raw = raw.astype(str).replace('nan', pd.NA).replace('None', pd.NA)
 
avg_lens = {c: raw[c].dropna().apply(lambda x: len(str(x))).mean() for c in raw.columns}
text_col = max(avg_lens, key=avg_lens.get)
 
# Build output
out = pd.DataFrame()
out['id'] = [f"R_{i+1:03d}" for i in range(len(raw))]
out['text'] = raw[text_col].fillna('').astype(str).str.strip()
 
# Map known columns by header name
def find_col(keywords):
for c in raw.columns:
if any(kw.lower() in str(c).lower() for kw in keywords):
return raw[c].fillna('Unknown').astype(str).str.strip()
return pd.Series(['Unknown'] * len(raw))
 
setting_raw = find_col(['practice setting','setting','institution','site','hospital'])
out['setting'] = setting_raw.apply(simplify_setting)
out['specialty'] = find_col(['specialty','specialt','spec','discipline'])
out['vendor'] = find_col(['cgp','vendor','lab','company','brand','tempus','fmi','caris','guardant'])
 
# Filter invalid responses
def is_valid(text):
t = str(text).strip()
if len(t) < 15: return False
if INVALID_RE.match(t): return False
return True
 
out['text_lower'] = out['text'].str.lower()
out = out[out['text'].apply(is_valid)].reset_index(drop=True)
out['id'] = [f"R_{i+1:03d}" for i in range(len(out))]
 
return out, None
 
# ─────────────────────────────────────────────────────────────────────────────
# HIGHLIGHT + QUOTE CARD
# ─────────────────────────────────────────────────────────────────────────────
def hl_text(text, focus=None):
result = text
order = (focus or []) + [t for t in THEMES if not focus or t not in focus]
pairs = []
for theme in order:
css = HL_MAP.get(theme, "hl-def")
for p in THEMES[theme]:
pairs.append((p, css, len(p)))
pairs.sort(key=lambda x: -x[2])
for p, css, _ in pairs:
result = re.sub(f'({re.escape(p)})', f'<span class="{css}">\\1</span>',
result, flags=re.IGNORECASE, count=3)
return result
 
def vendor_badge(vendor):
color = VENDOR_COLORS.get(vendor, "#3b6ef7")
return f'<span style="background:{color}22;color:{color};border:1px solid {color}44;padding:2px 10px;border-radius:99px;font-size:11px;font-weight:600">{vendor}</span>'
 
def setting_badge(setting):
color = "#fbbf24" if setting in ["Community","Private Practice"] else "#93c5fd"
css = "seg-a" if setting in ["Community","Private Practice"] else "seg-b"
return f'<span class="{css}">{setting}</span>'
 
def quote_card(row, focus=None):
h = hl_text(str(row.get("text", "")), focus)
vbadge = vendor_badge(str(row.get("vendor", "")))
sbadge = setting_badge(str(row.get("setting", "")))
tags = "".join(
f'<span class="tag" style="background:{TC.get(t,"#4a6080")}22;color:{TC.get(t,"#4a6080")};border:1px solid {TC.get(t,"#4a6080")}33">{t}</span>'
for t, pats in THEMES.items() if any(p in str(row.get("text_lower","")) for p in pats)
)
st.markdown(f"""<div class="fq-card">
<div class="fq-meta">📎 <b style="color:#e2ecf8">{row["id"]}</b> &nbsp;{vbadge}&nbsp;{sbadge}&nbsp;
<span style="color:#8a9ab5">{row.get("specialty","")}</span></div>
<div class="fq-text">{h}</div>
<div style="margin-top:8px;border-top:1px solid #1a2640;padding-top:6px">{tags}</div>
</div>""", unsafe_allow_html=True)
 
# ─────────────────────────────────────────────────────────────────────────────
# ANALYTICS
# ─────────────────────────────────────────────────────────────────────────────
def mtch(tl, pats): return any(p in tl for p in pats)
def t_counts(df): return {t: int(df["text_lower"].apply(lambda x: mtch(x,p)).sum()) for t,p in THEMES.items()}
 
def bar_html(label, count, total, color="#3b6ef7"):
pct = round(count/total*100) if total else 0
st.markdown(f"""<div class="bar-row">
<div class="bar-lbl">{label}</div>
<div class="bar-trk"><div class="bar-fill" style="width:{pct}%;background:{color}"></div></div>
<div class="bar-cnt">{count} <span style="color:#2a3a55">({pct}%)</span></div>
</div>""", unsafe_allow_html=True)
 
# ─────────────────────────────────────────────────────────────────────────────
# SENTIMENT ENGINE
# ─────────────────────────────────────────────────────────────────────────────
SENT_POSITIVE = [
"excellent","outstanding","exceptional","superb","fantastic","wonderful","amazing",
"great","good","very good","really good","pretty good","quite good",
"helpful","very helpful","extremely helpful","really helpful",
"responsive","very responsive","quick response","fast response",
"knowledgeable","very knowledgeable","well informed","expertise",
"easy","easy to use","easy to order","easy to access","straightforward",
"accurate","accuracy","reliable","reliability","precise","precision",
"fast","quick","rapid","timely","efficient","on time","quick turnaround",
"comprehensive","complete","thorough","detailed","broad","extensive",
"clear","clear report","easy to read","easy to interpret","well organized",
"actionable","clinically useful","clinically relevant","changes management",
"good support","strong support","excellent support","support was great",
"like","love","prefer","appreciate","impressed","satisfied","positive",
"best","top","leading","excellent company","solid company","good company",
"seamless","smooth","streamlined","integrated well","workflow",
"valuable","useful","informative","educational","insightful",
]
 
SENT_NEGATIVE = [
"poor","bad","terrible","awful","horrible","disappointing","disappointed",
"slow","too slow","slow turnaround","takes too long","takes forever",
"inaccurate","not accurate","errors","mistakes","wrong result","false",
"difficult","hard to","complicated","confusing","not easy","not user friendly",
"not helpful","unhelpful","not useful","limited support","poor support",
"expensive","costly","high cost","too expensive","cost prohibitive",
"not covered","no coverage","insurance issue","billing problem",
"unclear report","hard to interpret","confusing report","not clear",
"not actionable","not useful clinically","limited clinical utility",
"not responsive","slow to respond","hard to reach","no response",
"not comprehensive","limited panel","small panel","not enough genes",
"not integrated","no integration","no emr","manual process",
"not available","limited access","access issue","not reliable",
"concern","worried","hesitant","not sure","uncertain","doubt",
"never","don't use","do not use","stopped using","switched away",
"not the best","not great","mediocre","average at best",
]
 
NEGATION_PREFIXES = [
"no ","not ","never ","without ","don't have ","do not have ",
"doesn't have ","does not have ","haven't had ","have not had ",
"no significant ","no major ","no real ","haven't seen ",
"i haven't ","i have not ","we haven't ","we have not ",
"nothing ","absence of ","free of ","no longer ",
]
 
def classify_sentence(sentence):
sl = sentence.lower()
pos_hits, neg_hits = [], []
for sig in SENT_POSITIVE:
if sig in sl: pos_hits.append(sig)
for sig in SENT_NEGATIVE:
if sig in sl:
idx = sl.find(sig)
window = sl[max(0,idx-50):idx]
negated = any(neg in window for neg in NEGATION_PREFIXES)
if negated: pos_hits.append(f"no {sig}")
else: neg_hits.append(sig)
pos_hits = list(dict.fromkeys(pos_hits))[:4]
neg_hits = list(dict.fromkeys(neg_hits))[:4]
if pos_hits and neg_hits: s,conf = "MIXED","HIGH" if len(pos_hits)+len(neg_hits)>=3 else "MEDIUM"
elif pos_hits: s,conf = "POSITIVE","HIGH" if len(pos_hits)>=2 else "MEDIUM"
elif neg_hits: s,conf = "NEGATIVE","HIGH" if len(neg_hits)>=2 else "MEDIUM"
else: s,conf = "NEUTRAL","LOW"
return s, pos_hits, neg_hits, conf
 
def run_sentiment(theme, full_df):
pats = THEMES.get(theme, [theme.lower()]) if theme != "ALL" else []
rel_df = full_df[full_df["text_lower"].apply(lambda x: any(p in x for p in pats))] if pats else full_df
results = []
for _, row in rel_df.iterrows():
sents = [s.strip() for s in re.split(r'(?<=[.!?])\s+', str(row["text"])) if len(s.strip()) > 15]
if not sents: sents = [str(row["text"])]
for sent in sents:
sl = sent.lower()
theme_rel = True
if pats: theme_rel = any(p in sl for p in pats)
s, pos_t, neg_t, conf = classify_sentence(sent)
sent_themes = [t for t, tp in THEMES.items() if any(p in sl for p in tp)]
results.append({"id":row["id"],"setting":row["setting"],"specialty":row["specialty"],
"vendor":row["vendor"],"sentence":sent,"sentiment":s,
"pos_triggers":pos_t,"neg_triggers":neg_t,"confidence":conf,
"theme_relevant":theme_rel,"sentence_themes":sent_themes})
theme_sents = [r for r in results if r["theme_relevant"]]
T = len(theme_sents)
counts = {"POSITIVE":0,"NEGATIVE":0,"MIXED":0,"NEUTRAL":0}
for r in theme_sents: counts[r["sentiment"]] += 1
# Per-segment
seg_sent = {}
for col in ["setting","specialty","vendor"]:
seg_sent[col] = {}
for val in set(r[col] for r in theme_sents if r[col]):
vs = [r for r in theme_sents if r[col]==val]
vt = len(vs)
n_resp = len(set(r["id"] for r in vs))
seg_sent[col][val] = {"total":vt,"n_resp":n_resp,
"pos":sum(1 for r in vs if r["sentiment"]=="POSITIVE"),
"neg":sum(1 for r in vs if r["sentiment"]=="NEGATIVE"),
"mixed":sum(1 for r in vs if r["sentiment"]=="MIXED"),
"neutral":sum(1 for r in vs if r["sentiment"]=="NEUTRAL")}
from collections import Counter
all_pos = Counter(t for r in theme_sents for t in r["pos_triggers"])
all_neg = Counter(t for r in theme_sents for t in r["neg_triggers"])
return {"theme":theme,"n_respondents":len(rel_df),"total_sentences":T,"counts":counts,
"seg_sentiment":seg_sent,"all_results":theme_sents,
"top_pos":dict(all_pos.most_common(8)),"top_neg":dict(all_neg.most_common(8))}
 
# ─────────────────────────────────────────────────────────────────────────────
# DASHBOARD ENGINE
# ─────────────────────────────────────────────────────────────────────────────
def build_segment_breakdown(theme, full_df):
pats = THEMES.get(theme,[theme.lower()])
result = {}
for col in ["setting","specialty","vendor"]:
bd = {}
for val in full_df[col].dropna().unique():
seg = full_df[full_df[col]==val]
n_seg = len(seg)
n_match = int(seg["text_lower"].apply(lambda x: any(p in x for p in pats)).sum())
if n_seg > 0: bd[val] = {"n":n_match,"total":n_seg,"pct":round(n_match/n_seg*100)}
result[col] = bd
return result
 
def build_exec_summary(theme, theme_df, full_df, seg_data):
T = len(full_df); n = len(theme_df); pct = round(n/T*100) if T else 0
points = []
freq = "one of the most prominent" if pct>=50 else "a notable" if pct>=30 else "a less frequently cited"
points.append(f"**{n} of {T} respondents ({pct}%)** mentioned **{theme}**, making it {freq} topic across the dataset.")
s = seg_data.get("vendor",{})
if len(s)>=2:
items = sorted(s.items(), key=lambda x:-x[1]["pct"])
top,bot = items[0],items[-1]
gap = top[1]["pct"]-bot[1]["pct"]
if gap>=10:
points.append(f"**{top[0]}** respondents discuss {theme} more than **{bot[0]}** ({top[1]['pct']}% vs {bot[1]['pct']}%) — a **{gap}pp gap** suggesting vendor-specific differentiation on this topic.")
else:
points.append(f"Mentions are broadly consistent across CGP vendors — **{top[0]}**: {top[1]['pct']}% vs **{bot[0]}**: {bot[1]['pct']}% — no major vendor-driven divide.")
sp = seg_data.get("setting",{})
if len(sp)>=2:
items = sorted(sp.items(),key=lambda x:-x[1]["pct"])
top_sp = items[0]
others = "; ".join(f"{v}: {d['pct']}%" for v,d in items[1:3])
points.append(f"By practice setting, **{top_sp[0]}** most frequently discusses {theme} ({top_sp[1]['pct']}% of that group). {f'Other settings: {others}.' if others else ''}")
co = {}
for t,p2 in THEMES.items():
if t==theme: continue
c = int(theme_df["text_lower"].apply(lambda x: any(p in x for p in p2)).sum())
if c: co[t]=c
if co:
top3 = sorted(co.items(),key=lambda x:-x[1])[:3]
co_str = ", ".join(f"**{t}** ({c})" for t,c in top3)
points.append(f"{theme} most frequently co-occurs with {co_str} — suggesting it is part of a broader evaluation framework rather than a standalone consideration.")
solo = sum(1 for _,row in theme_df.iterrows() if sum(1 for t,p2 in THEMES.items() if t!=theme and any(p in row["text_lower"] for p in p2))<=1)
comp = sum(1 for _,row in theme_df.iterrows() if sum(1 for t,p2 in THEMES.items() if t!=theme and any(p in row["text_lower"] for p in p2))>=4)
if solo>comp:
points.append(f"**{theme} acts as a standalone evaluation criterion** — {solo} of {n} respondents ({round(solo/n*100) if n else 0}%) cited it without many other factors, indicating it independently influences CGP vendor selection.")
elif comp>solo:
points.append(f"**{theme} is a complex, multi-factor consideration** — {comp} of {n} respondents ({round(comp/n*100) if n else 0}%) discussed it alongside 4+ other topics, suggesting it is weighed as part of a broader vendor assessment.")
else:
points.append(f"**{theme} shows mixed complexity** — cited both independently ({solo}) and as part of multi-theme evaluations ({comp}).")
return points
 
def render_dashboard(theme, full_df):
pats = THEMES.get(theme,[theme.lower()])
theme_df = full_df[full_df["text_lower"].apply(lambda x: any(p in x for p in pats))]
T = len(full_df); n = len(theme_df); color = TC.get(theme,"#3b6ef7")
seg_data = build_segment_breakdown(theme,full_df)
 
st.markdown(f"""<div style="background:#0d1420;border:1px solid #1a2640;border-radius:14px;padding:20px 24px;margin-bottom:20px;border-left:4px solid {color}">
<div style="font-size:10px;letter-spacing:2px;text-transform:uppercase;color:{color};font-weight:600;margin-bottom:6px;font-family:'IBM Plex Mono',monospace">THEME DASHBOARD</div>
<div style="font-size:26px;font-weight:700;color:#e2ecf8;margin-bottom:4px">{theme}</div>
<div style="font-size:13px;color:#4a6080">{T} respondents (invalid responses excluded) · All data from uploaded Excel · No inference</div>
</div>""", unsafe_allow_html=True)
 
if n==0:
st.warning(f"No respondents mentioned {theme}.")
return
 
tab1, tab2, tab3 = st.tabs(["📋 Dashboard Summary","💬 Quotes & Evidence","⬇ Download"])
 
with tab1:
co = {}
for t,p2 in THEMES.items():
if t==theme: continue
c = int(theme_df["text_lower"].apply(lambda x: any(p in x for p in p2)).sum())
if c: co[t]=c
top_co = max(co,key=co.get) if co else "—"
solo = sum(1 for _,row in theme_df.iterrows() if sum(1 for t,p2 in THEMES.items() if t!=theme and any(p in row["text_lower"] for p in p2))<=1)
verdict = "Standalone" if solo>n//2 else "Complex"
vc = "#10b981" if verdict=="Standalone" else "#ef4444"
 
c1,c2,c3,c4 = st.columns(4)
c1.markdown(f'<div class="card" style="text-align:center;border-left:3px solid {color}"><div class="stat-num" style="color:{color}">{n}</div><div class="stat-lbl">respondents</div></div>',unsafe_allow_html=True)
c2.markdown(f'<div class="card" style="text-align:center"><div class="stat-num">{round(n/T*100)}%</div><div class="stat-lbl">of total sample</div></div>',unsafe_allow_html=True)
c3.markdown(f'<div class="card" style="text-align:center"><div class="stat-num" style="font-size:16px;line-height:1.4">{top_co}</div><div class="stat-lbl">top co-theme</div></div>',unsafe_allow_html=True)
c4.markdown(f'<div class="card" style="text-align:center"><div class="stat-num" style="color:{vc};font-size:20px">{verdict}</div><div class="stat-lbl">driver type</div></div>',unsafe_allow_html=True)
 
st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown('<div class="sec-lbl">5-POINT EXECUTIVE SUMMARY</div>', unsafe_allow_html=True)
st.markdown('<div style="font-size:11px;color:#4a6080;margin-bottom:14px">Synthesised from uploaded data only. Every point backed by counts.</div>', unsafe_allow_html=True)
for i,pt in enumerate(build_exec_summary(theme,theme_df,full_df,seg_data),1):
st.markdown(f"""<div style="display:flex;gap:14px;margin-bottom:14px;padding-bottom:14px;border-bottom:1px solid #1a2640;align-items:flex-start">
<div style="font-family:'IBM Plex Mono',monospace;font-size:18px;color:{color};font-weight:700;min-width:28px;line-height:1.4">{i}</div>
<div style="font-size:14px;color:#c8d4e8;line-height:1.75">{pt}</div>
</div>""", unsafe_allow_html=True)
st.markdown('</div>', unsafe_allow_html=True)
 
seg_cols_ui = st.columns(3)
for col_w,(seg_col,label) in zip(seg_cols_ui,[("setting","BY PRACTICE SETTING"),("specialty","BY SPECIALTY"),("vendor","BY CGP VENDOR")]):
bd = seg_data.get(seg_col,{})
if not bd: continue
with col_w:
st.markdown(f'<div class="card"><div class="sec-lbl">{label}</div>', unsafe_allow_html=True)
for val,dp in sorted(bd.items(),key=lambda x:-x[1]["pct"]):
vendor_c = VENDOR_COLORS.get(val, color)
bar_color = vendor_c if seg_col=="vendor" else color
st.markdown(f"""<div style="margin-bottom:10px">
<div style="display:flex;justify-content:space-between;font-size:12px;margin-bottom:3px">
<span style="color:#e2ecf8;font-weight:500">{val[:25]}</span>
<span style="font-family:'IBM Plex Mono',monospace;color:{bar_color}">{dp["n"]}/{dp["total"]} ({dp["pct"]}%)</span>
</div>
<div class="bar-trk"><div class="bar-fill" style="width:{dp["pct"]}%;background:{bar_color}"></div></div>
</div>""", unsafe_allow_html=True)
st.markdown('</div>', unsafe_allow_html=True)
 
st.markdown(f'<div class="card"><div class="sec-lbl">MOST ASSOCIATED THEMES</div>', unsafe_allow_html=True)
st.markdown(f'<div style="font-size:11px;color:#4a6080;margin-bottom:12px">Among {n} respondents who mentioned {theme}, these other themes also appeared.</div>', unsafe_allow_html=True)
for t_co,cnt_co in sorted(co.items(),key=lambda x:-x[1])[:10]:
if cnt_co: bar_html(t_co, cnt_co, n, TC.get(t_co,"#4a6080"))
st.markdown('</div>', unsafe_allow_html=True)
 
with tab2:
all_vendors = ["All"] + sorted(theme_df["vendor"].dropna().unique().tolist())
fv = st.selectbox("Filter by CGP Vendor", all_vendors, key=f"dash_v_{theme}")
disp = theme_df if fv=="All" else theme_df[theme_df["vendor"]==fv]
st.markdown(f'<div style="font-size:11px;color:#4a6080;margin-bottom:12px">Showing {len(disp)} respondents. Full text, all themes colour-highlighted.</div>', unsafe_allow_html=True)
for _,row in disp.iterrows():
quote_card(row, [theme])
 
with tab3:
export_rows = []
for _,row in theme_df.iterrows():
others = [t for t,p2 in THEMES.items() if t!=theme and any(p in row["text_lower"] for p in p2)]
export_rows.append({"ID":row["id"],"Practice_Setting":row["setting"],"Specialty":row["specialty"],
"CGP_Vendor":row["vendor"],f"Mentioned_{theme.replace(' ','_')}":"Yes",
"Other_Themes":", ".join(others[:8]),"N_Other_Themes":len(others),
"Driver_Type":"Standalone" if len(others)<=1 else "Complex" if len(others)>=4 else "Moderate",
"Full_Response":row["text"]})
dfe = pd.DataFrame(export_rows)
preview_cols = [c for c in ["ID","Practice_Setting","Specialty","CGP_Vendor","Driver_Type","N_Other_Themes"] if c in dfe.columns]
st.dataframe(dfe[preview_cols], hide_index=True, use_container_width=True)
c1,c2 = st.columns(2)
with c1:
st.download_button(f"⬇ {theme} full responses CSV", dfe.to_csv(index=False).encode(), f"{theme.replace('/','_').replace(' ','_')}_responses.csv","text/csv")
with c2:
sum_rows = [{"Segment_Type":lbl,"Segment_Value":val,"N_Mentioned":dp["n"],"Total_in_Group":dp["total"],"Pct_Mentioned":f"{dp['pct']}%"}
for seg_col,lbl in [("setting","Setting"),("specialty","Specialty"),("vendor","CGP Vendor")]
for val,dp in seg_data.get(seg_col,{}).items()]
st.download_button(f"⬇ {theme} segment summary CSV", pd.DataFrame(sum_rows).to_csv(index=False).encode(), f"{theme.replace('/','_').replace(' ','_')}_segments.csv","text/csv")
 
# ─────────────────────────────────────────────────────────────────────────────
# SENTIMENT VIEWS
# ─────────────────────────────────────────────────────────────────────────────
def render_theme_sentiment(theme, full_df):
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
with st.spinner(f"Running sentiment analysis for {theme}..."):
sa = run_sentiment(theme, full_df)
T_sents = sa["total_sentences"]
if T_sents==0:
st.warning(f"No sentences referencing {theme} found.")
return
color = TC.get(theme,"#3b6ef7")
counts = sa["counts"]
pos_n=counts["POSITIVE"]; neg_n=counts["NEGATIVE"]; mix_n=counts["MIXED"]; neu_n=counts["NEUTRAL"]
denom = pos_n+neg_n+neu_n
nss = round((pos_n-neg_n)/denom*100,1) if denom else 0
nss_color = "#10b981" if nss>10 else "#ef4444" if nss<-10 else "#f59e0b"
nss_label = "Net positive" if nss>10 else "Net negative" if nss<-10 else "Balanced"
 
st.markdown(f'''<div style="background:#0d1420;border:1px solid #1a2640;border-radius:12px;padding:16px 20px;margin-bottom:16px;border-left:4px solid {color}">
<div style="font-size:10px;letter-spacing:2px;text-transform:uppercase;color:{color};font-weight:600;margin-bottom:4px;font-family:'IBM Plex Mono',monospace">SENTIMENT ANALYSIS</div>
<div style="font-size:22px;font-weight:700;color:#e2ecf8">{theme}</div>
<div style="font-size:11px;color:#4a6080;margin-top:4px">{T_sents} sentences · {sa["n_respondents"]} respondents · {len(full_df)} total</div>
</div>''', unsafe_allow_html=True)
 
st.markdown(f'''<div class="card" style="border-left:4px solid {nss_color};margin-bottom:16px">
<div style="display:flex;align-items:center;gap:24px;flex-wrap:wrap">
<div style="text-align:center;min-width:110px">
<div style="font-size:10px;color:{nss_color};font-weight:600;letter-spacing:1.5px;text-transform:uppercase;margin-bottom:4px">Net Sentiment Score</div>
<div style="font-family:'IBM Plex Mono',monospace;font-size:44px;font-weight:700;color:{nss_color};line-height:1">{nss:+.1f}</div>
<div style="font-size:11px;color:{nss_color};margin-top:2px">{nss_label}</div>
</div>
<div style="flex:1;min-width:200px">
<div style="font-size:11px;color:#4a6080">Formula: (Positive − Negative) ÷ (Pos + Neg + Neutral) × 100<br>
({pos_n} − {neg_n}) ÷ ({pos_n} + {neg_n} + {neu_n}) × 100 = <b style="color:{nss_color}">{nss:+.1f}</b></div>
</div>
</div>
</div>''', unsafe_allow_html=True)
 
c1,c2,c3,c4 = st.columns(4)
for col,(label,count,col_c) in zip([c1,c2,c3,c4],[("Positive",pos_n,"#10b981"),("Negative",neg_n,"#ef4444"),("Mixed",mix_n,"#f59e0b"),("Neutral",neu_n,"#64748b")]):
pct = round(count/T_sents*100) if T_sents else 0
col.markdown(f'<div class="card" style="text-align:center;border-left:3px solid {col_c}"><div style="font-size:10px;color:{col_c};font-weight:600;margin-bottom:4px;text-transform:uppercase">{label}</div><div style="font-family:\'IBM Plex Mono\',monospace;font-size:28px;font-weight:600;color:{col_c}">{count}</div><div style="font-size:13px;color:#e2ecf8;margin-top:2px">{pct}%</div></div>',unsafe_allow_html=True)
 
pos_pct=round(pos_n/T_sents*100) if T_sents else 0
neg_pct=round(neg_n/T_sents*100) if T_sents else 0
mix_pct=round(mix_n/T_sents*100) if T_sents else 0
st.markdown(f'''<div style="margin:14px 0">
<div style="font-size:11px;color:#4a6080;margin-bottom:5px">Distribution — {T_sents} sentences</div>
<div style="display:flex;height:16px;border-radius:4px;overflow:hidden;gap:2px">
<div style="width:{pos_pct}%;background:#10b981"></div><div style="width:{neg_pct}%;background:#ef4444"></div>
<div style="width:{mix_pct}%;background:#f59e0b"></div><div style="flex:1;background:#1e293b"></div>
</div>
<div style="display:flex;gap:14px;margin-top:5px;font-size:11px">
<span style="color:#10b981">■ Positive {pos_pct}%</span><span style="color:#ef4444">■ Negative {neg_pct}%</span>
<span style="color:#f59e0b">■ Mixed {mix_pct}%</span><span style="color:#64748b">■ Neutral</span>
</div>
</div>''', unsafe_allow_html=True)
 
st.markdown('<div class="sec-lbl" style="margin-top:16px">SENTIMENT BY SEGMENT</div>', unsafe_allow_html=True)
for seg_key,seg_label in [("setting","Practice Setting"),("specialty","Specialty"),("vendor","CGP Vendor")]:
seg_data = sa["seg_sentiment"].get(seg_key,{})
if not seg_data: continue
st.markdown(f'''<div class="card" style="margin-bottom:14px"><div class="sec-lbl" style="margin-bottom:12px">{seg_label}</div>
<div style="display:grid;grid-template-columns:180px 1fr 60px 60px 60px 70px 70px 90px;font-size:10px;font-weight:600;color:#4a6080;letter-spacing:0.5px;text-transform:uppercase;padding:4px 0;border-bottom:1px solid #1a2640;margin-bottom:6px">
<span>Segment</span><span>Bar</span><span style="text-align:center">HCPs</span>
<span style="text-align:center">Pos</span><span style="text-align:center">Neg</span>
<span style="text-align:center">Neutral</span><span style="text-align:center">Sentences</span><span style="text-align:center">NSS</span>
</div>''', unsafe_allow_html=True)
for val,d in sorted(seg_data.items(),key=lambda x:-(x[1]["pos"]+x[1]["neg"])):
vt=d["total"]; pos2=d["pos"]; neg2=d["neg"]; neu2=d.get("neutral",0); n_resp2=d.get("n_resp",0)
p=round(pos2/vt*100) if vt else 0; n=round(neg2/vt*100) if vt else 0
denom2=pos2+neg2+neu2; sn=round((pos2-neg2)/denom2*100) if denom2 else 0
nc="#10b981" if sn>10 else "#ef4444" if sn<-10 else "#f59e0b"
nl="Net +" if sn>10 else "Net −" if sn<-10 else "Balanced"
warn="⚠" if n_resp2<5 or vt<10 else ""
vc2 = VENDOR_COLORS.get(val, color) if seg_key=="vendor" else color
pos_s=f"{p}%" if p>0 else "—"; neg_s=f"{n}%" if n>0 else "—"
st.markdown(f'''<div style="display:grid;grid-template-columns:180px 1fr 60px 60px 60px 70px 70px 90px;align-items:center;padding:7px 0;border-bottom:1px solid #0f1823">
<div style="font-size:12px;color:#e2ecf8;font-weight:500">{val[:25]}{warn}</div>
<div style="display:flex;height:10px;border-radius:3px;overflow:hidden;gap:2px;margin-right:8px;background:#1e293b">
<div style="width:{p}%;background:#10b981"></div><div style="width:{n}%;background:#ef4444"></div>
</div>
<div style="text-align:center"><div style="font-size:13px;color:#e2ecf8;font-weight:500">{n_resp2}</div><div style="font-size:9px;color:#4a6080">HCPs</div></div>
<div style="text-align:center;font-size:12px;color:#10b981">{pos_s}</div>
<div style="text-align:center;font-size:12px;color:{"#ef4444" if n>0 else "#4a6080"}">{neg_s}</div>
<div style="text-align:center;font-size:12px;color:#4a6080">{neu2}</div>
<div style="text-align:center"><div style="font-size:13px;color:#64748b">{vt}</div><div style="font-size:9px;color:#4a6080">sent.</div></div>
<div style="text-align:center"><span style="font-family:'IBM Plex Mono',monospace;font-size:14px;font-weight:700;color:{nc}">{sn:+d}</span><div style="font-size:9px;color:{nc}">{nl}</div></div>
</div>''', unsafe_allow_html=True)
# Calculation workings
st.markdown('<div style="margin-top:10px;border-top:1px solid #1a2640;padding-top:10px">', unsafe_allow_html=True)
st.markdown('<div style="font-size:10px;font-weight:600;color:#4a6080;letter-spacing:1px;text-transform:uppercase;margin-bottom:6px">NSS Calculation Workings</div>', unsafe_allow_html=True)
st.markdown('''<div style="display:grid;grid-template-columns:180px 50px 50px 50px 50px 1fr 80px;font-size:9px;font-weight:600;color:#4a6080;text-transform:uppercase;padding:4px 8px;background:#0a1220;border-radius:6px 6px 0 0">
<span>Segment</span><span style="text-align:center">Pos</span><span style="text-align:center">Neg</span><span style="text-align:center">Neutral</span><span style="text-align:center">Total</span><span style="text-align:center">Formula</span><span style="text-align:center">NSS</span>
</div>''', unsafe_allow_html=True)
small_segs = []
for val,d in sorted(seg_data.items(),key=lambda x:-(x[1]["pos"]+x[1]["neg"])):
pos2=d["pos"]; neg2=d["neg"]; neu2=d.get("neutral",0); denom2=pos2+neg2+neu2
sn=round((pos2-neg2)/denom2*100) if denom2 else 0
nc="#10b981" if sn>10 else "#ef4444" if sn<-10 else "#f59e0b"
formula=f"({pos2}−{neg2})÷({pos2}+{neg2}+{neu2})×100 = {pos2-neg2}÷{denom2}×100"
if d.get("n_resp",0)<5 or d["total"]<10: small_segs.append(val)
st.markdown(f'''<div style="display:grid;grid-template-columns:180px 50px 50px 50px 50px 1fr 80px;font-size:11px;align-items:center;padding:5px 8px;border-bottom:1px solid #0f1823;background:#0d1420">
<span style="color:#c8d4e8;font-weight:500">{val[:25]}</span>
<span style="text-align:center;color:#10b981;font-family:'IBM Plex Mono',monospace">{pos2}</span>
<span style="text-align:center;color:{"#ef4444" if neg2>0 else "#4a6080"};font-family:'IBM Plex Mono',monospace">{neg2}</span>
<span style="text-align:center;color:#4a6080;font-family:'IBM Plex Mono',monospace">{neu2}</span>
<span style="text-align:center;color:#64748b;font-family:'IBM Plex Mono',monospace">{denom2}</span>
<span style="color:#64748b;font-size:10px;font-family:'IBM Plex Mono',monospace">{formula}</span>
<span style="text-align:center;font-family:'IBM Plex Mono',monospace;font-weight:700;color:{nc}">{sn:+d}</span>
</div>''', unsafe_allow_html=True)
if small_segs:
st.markdown(f'<div style="font-size:10px;color:#f59e0b;padding:6px 8px;background:#1a1200;border-radius:0 0 6px 6px">⚠ Small sample: {", ".join(small_segs[:3])} — treat NSS directionally (n&lt;5 HCPs or &lt;10 sentences)</div>', unsafe_allow_html=True)
else:
st.markdown('<div style="height:4px;background:#0a1220;border-radius:0 0 6px 6px"></div>', unsafe_allow_html=True)
st.markdown('</div></div>', unsafe_allow_html=True)
 
st.markdown(f'''<div class="card" style="border-left:3px solid #3b6ef7;margin-top:4px">
<div class="sec-lbl">HOW NSS WAS CALCULATED</div>
<div style="font-size:12px;color:#c8d4e8;line-height:1.75">Each respondent's text split into sentences → sentences referencing <b>{theme}</b> identified → each sentence classified Positive / Negative / Mixed / Neutral using a pharma-tuned lexicon with negation detection → NSS = (Positive − Negative) ÷ (Pos + Neg + Neutral) × 100</div>
<div style="margin-top:8px;font-size:11px;color:#4a6080"><b style="color:#10b981">+10 to +100</b> = Net positive · <b style="color:#f59e0b">-10 to +10</b> = Balanced · <b style="color:#ef4444">-100 to -10</b> = Net negative</div>
</div>''', unsafe_allow_html=True)
 
st.markdown('<div class="sec-lbl" style="margin-top:20px">⬇ DOWNLOAD SENTIMENT DATA</div>', unsafe_allow_html=True)
if st.button(f"⬇ Generate & Download {theme} Sentiment Excel", key=f"dl_sent_{theme}"):
buf = _build_sentiment_excel(sa, theme)
st.download_button(f"⬇ Download {theme.replace('/','_')}_Sentiment.xlsx", buf, f"{theme.replace('/','_').replace(' ','_')}_Sentiment.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
 
def _build_sentiment_excel(sa, theme):
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
wb = Workbook()
def hc(ws,r,c,v,bg="0B1F3A",fg="FFFFFF",bold=True,sz=10):
cell=ws.cell(r,c,v); cell.font=Font(name="Arial",bold=bold,color=fg,size=sz)
cell.fill=PatternFill("solid",fgColor=bg); cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
def dc(ws,r,c,v,fg="1E293B",bold=False,sz=10,align="left",bg=None):
cell=ws.cell(r,c,v); cell.font=Font(name="Arial",bold=bold,color=fg,size=sz)
if bg: cell.fill=PatternFill("solid",fgColor=bg)
cell.alignment=Alignment(horizontal=align,vertical="center",wrap_text=True)
sc={"POSITIVE":"ECFDF5","NEGATIVE":"FEF2F2","MIXED":"FFFBEB","NEUTRAL":None}
sf={"POSITIVE":"065F46","NEGATIVE":"7F1D1D","MIXED":"92400E","NEUTRAL":"334155"}
ws1=wb.active; ws1.title="All Sentences"; ws1.sheet_view.showGridLines=False
for col,h in enumerate(["HCP_ID","Practice_Setting","Specialty","CGP_Vendor","Sentiment","Confidence","Positive_Triggers","Negative_Triggers","Full_Sentence"],1):
hc(ws1,1,col,h,"0D9488")
for ri,r in enumerate(sa["all_results"],2):
bg=sc.get(r["sentiment"]); fg2=sf.get(r["sentiment"],"334155")
for col,v in enumerate([r["id"],r["setting"],r["specialty"],r["vendor"],r["sentiment"],r["confidence"],", ".join(r["pos_triggers"]),", ".join(r["neg_triggers"]),r["sentence"]],1):
dc(ws1,ri,col,v,fg2 if col==5 else "1E293B",col==5,9,"left",bg if col==5 else None)
ws1.row_dimensions[ri].height=30
for col,w in zip(range(1,10),[10,16,16,16,12,11,28,22,55]):
ws1.column_dimensions[__import__("openpyxl").utils.get_column_letter(col)].width=w
ws2=wb.create_sheet("HCP NSS Summary"); ws2.sheet_view.showGridLines=False
hcp={}
for r in sa["all_results"]:
hid=r["id"]
if hid not in hcp: hcp[hid]={"id":hid,"setting":r["setting"],"specialty":r["specialty"],"vendor":r["vendor"],"pos":0,"neg":0,"mix":0,"neu":0}
hcp[hid][{"POSITIVE":"pos","NEGATIVE":"neg","MIXED":"mix","NEUTRAL":"neu"}.get(r["sentiment"],"neu")]+=1
for col,h in enumerate(["HCP_ID","Practice_Setting","Specialty","CGP_Vendor","Positive","Negative","Mixed","Neutral","NSS","NSS_Label"],1):
hc(ws2,1,col,h,"0D9488")
hcp_list=[]
for h in hcp.values():
d2=h["pos"]+h["neg"]+h["neu"]; hn=round((h["pos"]-h["neg"])/d2*100,1) if d2 else 0
hl="Net positive" if hn>10 else "Net negative" if hn<-10 else "Balanced"
hcp_list.append((h,hn,hl))
for ri,(h,hn,hl) in enumerate(sorted(hcp_list,key=lambda x:-x[1]),2):
bg2="ECFDF5" if hn>10 else "FEF2F2" if hn<-10 else "FFFBEB"
fg3="065F46" if hn>10 else "7F1D1D" if hn<-10 else "92400E"
for col,v in enumerate([h["id"],h["setting"],h["specialty"],h["vendor"],h["pos"],h["neg"],h.get("mix",0),h["neu"]],1):
dc(ws2,ri,col,v,"1E293B",False,10,"center")
dc(ws2,ri,9,f"{hn:+.1f}",fg3,True,11,"center",bg2); dc(ws2,ri,10,hl,fg3,False,10,"center",bg2)
ws2.row_dimensions[ri].height=20
for col,w in zip(range(1,11),[12,18,16,16,10,10,10,10,12,14]):
ws2.column_dimensions[__import__("openpyxl").utils.get_column_letter(col)].width=w
buf=__import__("io").BytesIO(); wb.save(buf); buf.seek(0)
return buf.getvalue()
 
def render_overall_sentiment(full_df):
st.markdown('''<div style="background:#0d1420;border:1px solid #1a2640;border-radius:14px;padding:18px 22px;margin-bottom:20px;border-left:4px solid #8b5cf6">
<div style="font-size:10px;letter-spacing:2px;text-transform:uppercase;color:#8b5cf6;font-weight:600;margin-bottom:6px;font-family:'IBM Plex Mono',monospace">SENTIMENT OVERVIEW</div>
<div style="font-size:22px;font-weight:700;color:#e2ecf8;margin-bottom:4px">All Themes — Net Sentiment Scores</div>
<div style="font-size:12px;color:#4a6080">NSS = (Positive − Negative) ÷ Total sentences × 100 · All 15 themes · By Setting, Specialty, CGP Vendor</div>
</div>''', unsafe_allow_html=True)
with st.spinner("Running sentiment across all themes..."):
all_results = {}
for theme in THEMES:
sa = run_sentiment(theme, full_df)
if sa["total_sentences"]>0: all_results[theme] = sa
if not all_results:
st.warning("No sentiment data found."); return
theme_nss = []
for theme,sa in all_results.items():
c=sa["counts"]; d=c["POSITIVE"]+c["NEGATIVE"]+c["NEUTRAL"]
nss=round((c["POSITIVE"]-c["NEGATIVE"])/d*100) if d else 0
theme_nss.append((theme,nss,c["POSITIVE"],c["NEGATIVE"],c["NEUTRAL"],c["MIXED"],sa["total_sentences"],sa["n_respondents"]))
theme_nss.sort(key=lambda x:-x[1])
st.markdown('<div class="sec-lbl">ALL THEMES — RANKED BY NSS</div>', unsafe_allow_html=True)
st.markdown('''<div style="background:#0d1420;border:1px solid #1a2640;border-radius:10px;overflow:hidden;margin-bottom:20px">
<div style="display:grid;grid-template-columns:200px 1fr 60px 60px 60px 70px 70px 90px;font-size:10px;font-weight:600;color:#4a6080;letter-spacing:0.5px;text-transform:uppercase;padding:8px 14px;background:#0a1220;border-bottom:1px solid #1a2640">
<span>Theme</span><span>NSS bar</span><span style="text-align:center">HCPs</span><span style="text-align:center">Pos</span><span style="text-align:center">Neg</span><span style="text-align:center">Neutral</span><span style="text-align:center">Sentences</span><span style="text-align:center">NSS</span>
</div>''', unsafe_allow_html=True)
for theme,nss,pos,neg,neu,mix,total_s,n_resp in theme_nss:
color=TC.get(theme,"#4a6080"); nss_c="#10b981" if nss>10 else "#ef4444" if nss<-10 else "#f59e0b"
bar_w=min(abs(nss),100)
st.markdown(f'''<div style="display:grid;grid-template-columns:200px 1fr 60px 60px 60px 70px 70px 90px;align-items:center;padding:7px 14px;border-bottom:1px solid #0f1823">
<div style="font-size:12px;color:#e2ecf8;font-weight:500"><span style="display:inline-block;width:8px;height:8px;border-radius:50%;background:{color};margin-right:6px"></span>{theme}</div>
<div style="display:flex;height:10px;border-radius:3px;overflow:hidden;margin-right:8px;background:#1e293b"><div style="width:{bar_w}%;background:{nss_c};border-radius:3px"></div></div>
<div style="text-align:center;font-size:12px;color:#e2ecf8;font-weight:500">{n_resp}</div>
<div style="text-align:center;font-size:12px;color:#10b981">{pos}</div>
<div style="text-align:center;font-size:12px;color:{"#ef4444" if neg>0 else "#4a6080"}">{neg if neg>0 else "—"}</div>
<div style="text-align:center;font-size:12px;color:#4a6080">{neu}</div>
<div style="text-align:center;font-size:12px;color:#64748b">{total_s}</div>
<div style="text-align:center"><span style="font-family:'IBM Plex Mono',monospace;font-size:14px;font-weight:700;color:{nss_c}">{nss:+d}</span><div style="font-size:9px;color:{nss_c}">{"Net +" if nss>10 else "Net −" if nss<-10 else "Balanced"}</div></div>
</div>''', unsafe_allow_html=True)
st.markdown('</div>', unsafe_allow_html=True)
st.markdown('<div class="sec-lbl" style="margin-top:20px">NSS BY SEGMENT — ALL THEMES</div>', unsafe_allow_html=True)
for seg_key,seg_label in [("setting","Practice Setting"),("specialty","Specialty"),("vendor","CGP Vendor")]:
seg_vals = sorted(full_df[seg_key].dropna().unique().tolist())
if not seg_vals: continue
st.markdown(f'<div class="card" style="margin-bottom:14px"><div class="sec-lbl" style="margin-bottom:14px">{seg_label}</div>', unsafe_allow_html=True)
ncols = len(seg_vals); col_w = f"200px " + " ".join(["1fr"]*ncols)
header = f'<div style="display:grid;grid-template-columns:{col_w};font-size:10px;font-weight:600;color:#4a6080;letter-spacing:0.5px;text-transform:uppercase;padding:6px 8px;background:#0a1220;border-radius:6px;margin-bottom:4px"><span>Theme</span>'
for v in seg_vals: header += f'<span style="text-align:center">{v[:14]}</span>'
header += '</div>'; st.markdown(header, unsafe_allow_html=True)
for theme,*_ in theme_nss:
sa2 = all_results.get(theme)
if not sa2: continue
seg_data2 = sa2["seg_sentiment"].get(seg_key,{})
row_html = f'<div style="display:grid;grid-template-columns:{col_w};align-items:center;padding:5px 8px;border-bottom:1px solid #0f1823"><span style="font-size:11px;color:#e2ecf8">{theme}</span>'
for v in seg_vals:
d = seg_data2.get(v)
if d:
d2=d["pos"]+d["neg"]+d.get("neutral",0); sn=round((d["pos"]-d["neg"])/d2*100) if d2 else 0
nc="#10b981" if sn>10 else "#ef4444" if sn<-10 else "#f59e0b"
warn="⚠" if d.get("n_resp",0)<5 else ""
row_html += f'<div style="text-align:center"><span style="font-family:\'IBM Plex Mono\',monospace;font-size:12px;font-weight:700;color:{nc}">{sn:+d}{warn}</span></div>'
else: row_html += '<div style="text-align:center;color:#2a3a55">—</div>'
row_html += '</div>'; st.markdown(row_html, unsafe_allow_html=True)
st.markdown('</div>', unsafe_allow_html=True)
 
# ─────────────────────────────────────────────────────────────────────────────
# COMPARISON ENGINE
# ─────────────────────────────────────────────────────────────────────────────
def render_comparison(full_df):
st.markdown('''<div style="background:#0d1420;border:1px solid #1a2640;border-radius:14px;padding:18px 22px;margin-bottom:20px;border-left:4px solid #f59e0b">
<div style="font-size:10px;letter-spacing:2px;text-transform:uppercase;color:#f59e0b;font-weight:600;margin-bottom:6px;font-family:'IBM Plex Mono',monospace">COMPARISON VIEW</div>
<div style="font-size:22px;font-weight:700;color:#e2ecf8;margin-bottom:4px">Segment Comparison — All Themes</div>
<div style="font-size:12px;color:#4a6080">Compare % mentioning each theme by Practice Setting, Specialty, or CGP Vendor</div>
</div>''', unsafe_allow_html=True)
 
comp_by = st.selectbox("Compare by", ["CGP Vendor","Practice Setting","Specialty"], key="comp_by")
seg_col_map = {"CGP Vendor":"vendor","Practice Setting":"setting","Specialty":"specialty"}
seg_col = seg_col_map[comp_by]
seg_vals = sorted(full_df[seg_col].dropna().unique().tolist())
 
# Build comparison table
T = len(full_df)
rows = []
for theme, pats in THEMES.items():
row = {"Theme": theme}
for val in seg_vals:
seg = full_df[full_df[seg_col]==val]
n_seg = len(seg)
n_match = int(seg["text_lower"].apply(lambda x: any(p in x for p in pats)).sum())
pct = round(n_match/n_seg*100) if n_seg else 0
row[val] = pct
row[f"{val}_n"] = n_match
rows.append(row)
 
# Sort by variance across segments
rows.sort(key=lambda r: -max(r[v] for v in seg_vals if v in r))
 
# Header
ncols = len(seg_vals)
col_w = "200px " + " ".join(["1fr"]*ncols)
counts_header = {v: len(full_df[full_df[seg_col]==v]) for v in seg_vals}
vc_map = VENDOR_COLORS if comp_by=="CGP Vendor" else {}
 
st.markdown('<div style="background:#0d1420;border:1px solid #1a2640;border-radius:10px;overflow:hidden;margin-bottom:20px">', unsafe_allow_html=True)
header = f'<div style="display:grid;grid-template-columns:{col_w};font-size:10px;font-weight:600;color:#4a6080;letter-spacing:0.5px;text-transform:uppercase;padding:8px 14px;background:#0a1220;border-bottom:1px solid #1a2640"><span>Theme</span>'
for v in seg_vals:
vc = vc_map.get(v,"#3b6ef7")
header += f'<span style="text-align:center;color:{vc}">{v[:14]}<br><span style="font-size:9px;color:#4a6080">n={counts_header[v]}</span></span>'
header += '</div>'; st.markdown(header, unsafe_allow_html=True)
 
for row in rows:
theme = row["Theme"]; color = TC.get(theme,"#4a6080")
max_val = max(row.get(v,0) for v in seg_vals)
row_html = f'<div style="display:grid;grid-template-columns:{col_w};align-items:center;padding:7px 14px;border-bottom:1px solid #0f1823"><div style="font-size:12px;color:#e2ecf8;font-weight:500"><span style="display:inline-block;width:8px;height:8px;border-radius:50%;background:{color};margin-right:6px"></span>{theme}</div>'
for v in seg_vals:
pct = row.get(v,0); n = row.get(f"{v}_n",0)
is_max = pct==max_val and pct>0
vc = vc_map.get(v, color)
bar_color = vc if comp_by=="CGP Vendor" else color
bg_hl = f"background:{bar_color}15;" if is_max else ""
row_html += f'<div style="text-align:center;{bg_hl}padding:3px"><div style="font-family:\'IBM Plex Mono\',monospace;font-size:{"14px" if is_max else "12px"};font-weight:{"700" if is_max else "400"};color:{bar_color if is_max else "#64748b"}">{pct}%</div><div style="font-size:9px;color:#4a6080">{n} HCPs</div></div>'
row_html += '</div>'; st.markdown(row_html, unsafe_allow_html=True)
st.markdown('</div>', unsafe_allow_html=True)
 
# Download
export_data = []
for row in rows:
er = {"Theme": row["Theme"]}
for v in seg_vals:
er[f"{v} (%)"] = row.get(v,0)
er[f"{v} (n)"] = row.get(f"{v}_n",0)
export_data.append(er)
dfe = pd.DataFrame(export_data)
st.download_button("⬇ Download comparison CSV", dfe.to_csv(index=False).encode(), f"comparison_{comp_by.replace(' ','_')}.csv","text/csv")
 
# ─────────────────────────────────────────────────────────────────────────────
# QUESTION BANK — CGP/FMI specific
# ─────────────────────────────────────────────────────────────────────────────
QB = {
"📊 Frequency": [
"How many HCPs mentioned turnaround time?",
"How many mentioned EMR/EHR integration?",
"How many cited report clarity as a factor?",
"How often was test accuracy mentioned?",
"How many mentioned clinical trial matching?",
"How many discussed patient financial support?",
"How often was panel size mentioned?",
"How many cited physician support services?",
"How often was clinical utility discussed?",
"How many mentioned ease of ordering?",
],
"⚖️ Comparison": [
"FMI vs Tempus — what's different in feedback?",
"Caris vs Guardant — which gets more positive mentions?",
"Academic vs community on turnaround time",
"FMI vs Caris on report clarity",
"Tempus vs Guardant on EMR integration",
"Academic vs community on clinical trial matching",
"FMI vs Tempus on physician support services",
"MedOnc vs HemeOnc — what themes differ?",
"All four vendors on test accuracy",
"Academic vs community on patient financial support",
],
"🎯 Drivers": [
"What is the top driver for positive rep ratings?",
"What are the main reasons HCPs prefer FMI?",
"What drives positive feedback for Tempus?",
"What are the key evaluation criteria for CGP vendors?",
"What is most commonly cited as a differentiator?",
"What are the main barriers to switching vendors?",
"What drives preference for academic HCPs?",
"What are the top reasons for community HCP preferences?",
"What is the primary driver for Caris selections?",
"What drives Guardant Health preference?",
],
"🔗 Co-occurrence": [
"Did turnaround time and report clarity go together?",
"Did EMR integration and ease of ordering co-occur?",
"When clinical utility was mentioned, was accuracy also discussed?",
"Did physician support and clinical trial matching go together?",
"Did panel size and clinical utility co-occur?",
"When TAT was mentioned, was report quality also discussed?",
"Did patient financial support and ease of ordering go together?",
"Did research support and aggregated data co-occur?",
"When reflex testing was mentioned, what else came up?",
"Did EMR integration and ease of ordering go hand in hand?",
],
"🧩 Theme Clustering": [
"What other themes travel with turnaround time?",
"What clusters around report clarity?",
"What travels with EMR/EHR integration mentions?",
"What other topics come up with clinical utility?",
"What clusters around physician support services?",
"What travels with test accuracy mentions?",
"What other factors come up with panel size?",
"What clusters around clinical trial matching?",
"What travels with patient financial support mentions?",
"What other themes appear with ease of ordering?",
],
"🎯 Driver Complexity": [
"Was turnaround time a standalone driver or always complex?",
"Was report clarity a simple or entangled factor?",
"Was EMR integration a standalone consideration?",
"Was test accuracy a simple or complex evaluation criterion?",
"Was clinical utility standalone or bundled with other themes?",
"Was physician support a standalone driver?",
"Was panel size a simple or complex consideration?",
"Was ease of ordering standalone or always mentioned with others?",
],
"💬 Full Responses": [
"Show full responses mentioning turnaround time",
"Give me full quotes about report clarity",
"Show full responses about EMR integration",
"Give me full quotes about clinical trial matching",
"Show full responses about patient financial support",
"Show full responses from academic HCPs",
"Give me full quotes about physician support services",
"Show full responses from FMI respondents",
"Show full responses about test accuracy",
"Give me full quotes about ease of ordering",
],
"🏥 Vendor-Specific": [
"What do FMI respondents say most often?",
"What are the top themes for Tempus?",
"What does Caris Life Sciences feedback focus on?",
"What do Guardant Health respondents emphasise?",
"Show full responses from Caris respondents",
"What are the key differences between FMI and Tempus feedback?",
],
}
 
# ─────────────────────────────────────────────────────────────────────────────
# INTENT + ANSWER ENGINE
# ─────────────────────────────────────────────────────────────────────────────
def get_topics(q):
ql=q.lower()
return list(dict.fromkeys([t for t,pats in THEMES.items() if any(p in ql for p in pats) or t.lower() in ql]))
 
def get_intent(q):
ql=q.lower()
tpcs=get_topics(q)
if len(tpcs)>=2 and any(w in ql for w in ["hand in hand","together","alongside","linked","co-occur","both","went together"]): return "co_occur"
if any(w in ql for w in ["what else","other drivers","other themes","tagged to","cluster","travel with","come with","what other","drivers tagged"]): return "cluster"
if any(w in ql for w in ["standalone","straightforward","only driver","simple driver","complex","entangled","on its own","always with","was it just"]): return "complexity"
rules = {
"comparison":[r'(fmi|tempus|caris|guardant|academic|community|medonc|heme|setting|specialty|vendor).*(vs|versus|compare|differ|more|less|between|than)|(vs|versus|compare|differ|difference).*(fmi|tempus|caris|guardant|academic|community|setting|specialty|vendor)'],
"frequency": [r'how many|how often|what (percentage|proportion|%)|how frequent|number of|count'],
"driver": [r'(top|main|key|primary|most common|biggest|major|most important).*(reason|driver|factor|barrier|concern)|(most tied|most associated|which driver|what drives|why do)'],
"quotes": [r'(quote|verbatim|exact words|what.*(say|said)|show me.*(quote|example|response)|give me.*(quote|full response|full quote)|full response|full transcript)'],
"vendor": [r'(fmi|tempus|caris|guardant).*(say|mention|focus|rate|prefer|feedback|response)|(what do|show me).*(fmi|tempus|caris|guardant)'],
"barrier": [r'(barrier|obstacle|challenge|concern|hesitat|reluctan).*(vendor|switch|adopt|use|order)|what stops|what prevents'],
}
sc=defaultdict(int)
for i,pats in rules.items():
for p in pats:
if re.search(p,ql): sc[i]+=1
return max(sc,key=sc.get) if sc else "content"
 
def is_comp(q):
ql=q.lower()
sw=["fmi","tempus","caris","guardant","academic","community","medonc","heme","setting","specialty","vendor","office","private"]
cw=["vs","versus","compare","difference","different","more","less","between","than","higher","lower"]
return any(w in ql for w in sw) and any(w in ql for w in cw)
 
def detect_segs(q, df):
ql=q.lower()
for col in ["vendor","setting","specialty"]:
vals=df[col].unique().tolist()
hits=[v for v in vals if v.lower() in ql or v.lower().replace(" ","") in ql.replace(" ","")]
if len(hits)>=2: return (col,hits[0]),(col,hits[1])
if len(hits)==1:
others=[v for v in vals if v!=hits[0]]
if others: return (col,hits[0]),(col,others[0])
# Default vendor comparison
vendors=df["vendor"].unique().tolist()
if len(vendors)>=2: return ("vendor",vendors[0]),("vendor",vendors[1])
return None,None
 
def answer(q, adf, fdf):
T=len(fdf); itn=get_intent(q); tpcs=get_topics(q); comp=is_comp(q)
qpats=[]
for t in tpcs: qpats+=THEMES.get(t,[])
if not qpats: qpats=[w for w in q.lower().split() if len(w)>4]
res={"intent":itn,"topics":tpcs,"is_comp":comp,"n":len(adf),"T":T,"summary":"","chart":{},"rows":[],"comp":None,"export":[]}
 
if comp:
sa,sb=detect_segs(q,fdf)
if sa and sb:
ca,va=sa; cb,vb=sb
dfa=fdf[fdf[ca]==va]; dfb=fdf[fdf[cb]==vb]; na,nb=len(dfa),len(dfb)
tca,tcb=t_counts(dfa),t_counts(dfb)
rows=[]
for t in THEMES:
a,b=tca.get(t,0),tcb.get(t,0)
pa=round(a/na*100) if na else 0; pb=round(b/nb*100) if nb else 0
rows.append({"Theme":t,f"{va}(n={na})":f"{a}({pa}%)",f"{vb}(n={nb})":f"{b}({pb}%)","D":f"+{pa-pb}%" if pa>pb else f"{pa-pb}%","_d":pa-pb,"_pa":pa,"_pb":pb,"_a":a,"_b":b})
rows=([r for r in rows if r["Theme"] in tpcs] or rows) if tpcs else sorted(rows,key=lambda x:abs(x["_d"]),reverse=True)[:10]
top=rows[0] if rows else {}
if top:
w=va if top["_d"]>0 else vb; l=vb if top["_d"]>0 else va
res["summary"]=f"**{w}** mentions **{top['Theme']}** more than **{l}** by **{abs(top['_d'])}pp** ({top['_pa']}% vs {top['_pb']}%)."
res["comp"]={"rows":rows,"va":va,"vb":vb,"na":na,"nb":nb,"dfa":dfa,"dfb":dfb,"focus":qpats}
res["intent"]="comparison"; return res
 
mdf=adf[adf["text_lower"].apply(lambda x: any(p in x for p in qpats))] if qpats else adf
nm=len(mdf); pct=round(nm/T*100) if T else 0; tstr=" + ".join(tpcs) if tpcs else "this topic"
res["n"]=nm; res["rows"]=mdf.to_dict("records")
res["export"]=[{"ID":r["id"],"Practice_Setting":r["setting"],"Specialty":r["specialty"],"CGP_Vendor":r["vendor"],"Full_Response":r["text"]} for r in res["rows"]]
 
if itn in ["frequency","content"]:
res["summary"]=f"**{nm} of {T} respondents ({pct}%)** mentioned {tstr}."
res["chart"]={"Associated Themes":dict(sorted({t:v for t,v in t_counts(mdf).items() if v>0}.items(),key=lambda x:-x[1])[:10])}
elif itn=="driver":
top_themes=sorted(t_counts(mdf).items(),key=lambda x:-x[1])[:8]
res["summary"]=f"Top themes for **{tstr}**: **{top_themes[0][0]}** cited by {top_themes[0][1]} respondents." if top_themes else f"{nm} respondents matched."
res["chart"]={"Theme Frequency":dict(top_themes)}
elif itn in ["quotes","vendor"]:
res["summary"]=f"**{nm} respondents** discussed {tstr}. Showing full responses below."
elif itn=="barrier":
res["summary"]=f"**{nm} respondents** mentioned barriers around {tstr}."
res["chart"]={"Associated Themes":dict(sorted({t:v for t,v in t_counts(mdf).items() if v>0}.items(),key=lambda x:-x[1])[:8])}
else:
res["summary"]=f"**{nm} of {T} respondents ({pct}%)** discussed {tstr}."
res["chart"]={"Themes Found":dict(sorted({t:v for t,v in t_counts(mdf).items() if v>0}.items(),key=lambda x:-x[1])[:10])}
return res
 
# ─────────────────────────────────────────────────────────────────────────────
# APP
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""<div style="margin-bottom:20px">
<div style="font-family:'IBM Plex Mono',monospace;font-size:10px;color:#3b6ef7;letter-spacing:2px;margin-bottom:6px">CGP VENDOR INSIGHT ENGINE · UPLOAD YOUR DATA TO BEGIN</div>
<h1 style="font-size:30px;margin:0;color:#e2ecf8;font-weight:600">CGP Insight Engine</h1>
</div>""", unsafe_allow_html=True)
 
# ── SIDEBAR ───────────────────────────────────────────────────────────────────
with st.sidebar:
st.markdown('<div class="sec-lbl">📂 UPLOAD DATA</div>', unsafe_allow_html=True)
st.markdown('<div style="font-size:11px;color:#4a6080;margin-bottom:10px;line-height:1.6">Upload your CGP qual Excel.<br><span style="color:#8a9ab5">Required:</span> verbatim response column<br><span style="color:#8a9ab5">Optional:</span> Practice Setting · Specialty · CGP Vendor<br><span style="color:#2a3a55">Invalid/short responses auto-excluded.</span></div>', unsafe_allow_html=True)
uploaded = st.file_uploader("Upload Excel", type=["xlsx","xls"], label_visibility="collapsed")
st.markdown("<hr>", unsafe_allow_html=True)
 
if not uploaded:
st.markdown('<div style="text-align:center;padding:20px;color:#4a6080;font-size:12px">Upload an Excel file to begin.</div>', unsafe_allow_html=True)
st.stop()
 
full_df, err = load_excel(uploaded)
if err or full_df is None:
st.error(f"⚠️ {err}")
st.stop()
 
TOTAL = len(full_df)
st.markdown(f'<div style="background:#0a1f14;border:1px solid #10b98133;border-radius:10px;padding:12px 16px;margin-bottom:14px"><div style="font-size:11px;font-weight:600;color:#34d399;margin-bottom:3px">✅ Data loaded</div><div style="font-size:11px;color:#4a6080">{TOTAL} valid respondents · {uploaded.name}</div></div>', unsafe_allow_html=True)
 
# Filters
st.markdown('<div class="sec-lbl">SEGMENT FILTERS</div>', unsafe_allow_html=True)
settings = ["All"] + sorted(full_df["setting"].dropna().unique().tolist())
specialties = ["All"] + sorted(full_df["specialty"].dropna().unique().tolist())
vendors = ["All"] + sorted(full_df["vendor"].dropna().unique().tolist())
fs = st.selectbox("Practice Setting", settings)
fsp = st.selectbox("Specialty", specialties)
fv = st.selectbox("CGP Vendor", vendors)
 
fdf = full_df.copy()
if fs != "All": fdf = fdf[fdf["setting"] ==fs]
if fsp != "All": fdf = fdf[fdf["specialty"]==fsp]
if fv != "All": fdf = fdf[fdf["vendor"] ==fv]
fn=len(fdf)
vc=fdf["vendor"].value_counts().to_dict()
vh=" · ".join(f'<span style="color:{VENDOR_COLORS.get(k,"#4a6080")}">{k}:{v}</span>' for k,v in vc.items())
st.markdown(f'<div class="card" style="text-align:center;margin-top:6px"><div class="stat-num">{fn}</div><div class="stat-lbl">in view</div><div style="font-size:11px;margin-top:6px">{vh}</div></div>', unsafe_allow_html=True)
st.markdown("<hr>", unsafe_allow_html=True)
 
# 3 tabs
side_tab1, side_tab2, side_tab3, side_tab4 = st.tabs(["📊 Dashboard","❓ Questions","📊 Compare","🎭 Sentiment"])
 
with side_tab1:
st.markdown('<div style="font-size:11px;color:#4a6080;margin-bottom:10px;margin-top:8px">Click any theme for the full dashboard.</div>', unsafe_allow_html=True)
for theme_name in list(THEMES.keys()):
if st.button(f"● {theme_name}", key=f"dash_{theme_name}"):
st.session_state["dashboard_theme"]=theme_name; st.session_state["q"]=""; st.session_state["sent_theme"]=""; st.session_state["sent_overall"]=False; st.session_state["show_comp"]=False; st.rerun()
 
with side_tab2:
st.markdown('<div style="font-size:11px;color:#4a6080;margin-bottom:8px;margin-top:8px">Click to run instantly</div>', unsafe_allow_html=True)
for ci, (cat,qs) in enumerate(QB.items()):
with st.expander(cat, expanded=False):
for qi, q in enumerate(qs):
if st.button(q, key=f"qb_{ci}_{qi}"):
st.session_state["q"]=q; st.session_state["dashboard_theme"]=""; st.session_state["sent_theme"]=""; st.session_state["sent_overall"]=False; st.session_state["show_comp"]=False; st.rerun()
 
with side_tab3:
st.markdown('<div style="font-size:11px;color:#4a6080;margin-bottom:10px;margin-top:8px">Side-by-side % comparison across all vendors, settings, and specialties.</div>', unsafe_allow_html=True)
if st.button("🌐 Open Comparison View", key="open_comp"):
st.session_state["show_comp"]=True; st.session_state["dashboard_theme"]=""; st.session_state["q"]=""; st.session_state["sent_theme"]=""; st.session_state["sent_overall"]=False; st.rerun()
 
with side_tab4:
st.markdown('<div style="font-size:11px;color:#4a6080;margin-bottom:10px;margin-top:8px">NSS scores and segment breakdown. Quotes downloadable as Excel.</div>', unsafe_allow_html=True)
if st.button("🌐 All themes overview", key="sent_overall_btn"):
st.session_state["sent_overall"]=True; st.session_state["sent_theme"]=""; st.session_state["dashboard_theme"]=""; st.session_state["q"]=""; st.session_state["show_comp"]=False; st.rerun()
st.markdown("<hr>", unsafe_allow_html=True)
for theme_name in list(THEMES.keys()):
if st.button(f"● {theme_name}", key=f"sent_{theme_name}"):
st.session_state["sent_theme"]=theme_name; st.session_state["sent_overall"]=False; st.session_state["dashboard_theme"]=""; st.session_state["q"]=""; st.session_state["show_comp"]=False; st.rerun()
 
# ── MAIN ──────────────────────────────────────────────────────────────────────
for k,v in [("q",""),("dashboard_theme",""),("sent_theme",""),("sent_overall",False),("show_comp",False)]:
if k not in st.session_state: st.session_state[k]=v
 
def back_btn():
if st.button("← Back", key="back_btn"):
for k in ["dashboard_theme","sent_theme","sent_overall","show_comp","q"]:
st.session_state[k]="" if k not in [True,False] else False
st.session_state["sent_overall"]=False; st.session_state["show_comp"]=False
st.rerun()
 
if st.session_state.get("sent_overall"):
back_btn(); render_overall_sentiment(full_df); st.stop()
if st.session_state.get("sent_theme"):
back_btn(); render_theme_sentiment(st.session_state["sent_theme"], full_df); st.stop()
if st.session_state.get("show_comp"):
back_btn(); render_comparison(full_df); st.stop()
if st.session_state.get("dashboard_theme") and not st.session_state.get("q"):
back_btn(); render_dashboard(st.session_state["dashboard_theme"], full_df); st.stop()
 
# Search area
st.markdown('<div style="font-size:12px;color:#8a9ab5;margin-bottom:6px;font-weight:500">🔍 <b style="color:#e2ecf8">Ask a question</b> — or click from sidebar</div>', unsafe_allow_html=True)
query = st.text_input(label="q", label_visibility="collapsed",
placeholder='e.g. "FMI vs Tempus on turnaround time" · "How many mentioned EMR integration?" · "What travels with report clarity?"',
value=st.session_state.get("q",""), key="main")
 
st.markdown('<div style="font-size:10px;color:#4a6080;margin:6px 0 4px;letter-spacing:1px;text-transform:uppercase">Quick questions:</div>', unsafe_allow_html=True)
chip_cols=st.columns(4)
CHIPS=["FMI vs Tempus on turnaround time","How many mentioned report clarity?","What travels with EMR integration?","Academic vs community on TAT","Caris vs Guardant feedback","Was TAT standalone or complex?","Show full responses about accuracy","What drives positive Tempus ratings?"]
for i,chip in enumerate(CHIPS):
if chip_cols[i%4].button(chip,key=f"chip_{i}"):
st.session_state["q"]=chip; st.rerun()
 
st.markdown("""<div style="display:flex;flex-wrap:wrap;gap:8px;margin:10px 0 4px;align-items:center">
<span style="font-size:10px;color:#4a6080">Highlights:</span>
<span class="hl-testing" style="font-style:normal">testing modality</span>
<span class="hl-support" style="font-style:normal">physician support</span>
<span class="hl-accuracy" style="font-style:normal">test accuracy</span>
<span class="hl-report" style="font-style:normal">report clarity</span>
<span class="hl-cost" style="font-style:normal">patient financial</span>
<span class="hl-tat" style="font-style:normal">turnaround time</span>
<span class="hl-emr" style="font-style:normal">EMR/EHR</span>
</div>""", unsafe_allow_html=True)
 
if query and query.strip():
with st.spinner(""):
r = answer(query, fdf, full_df)
st.markdown("<hr>", unsafe_allow_html=True)
bh=f'<span class="ibadge">{r["intent"]}</span>'
for t in r["topics"]: bh+=f'<span class="tbadge">{t}</span>'
if fv!="All": bh+=f'<span class="ibadge" style="color:{VENDOR_COLORS.get(fv,"#f59e0b")}">{fv}</span>'
st.markdown(bh+"<br><br>",unsafe_allow_html=True)
st.markdown(f'<div class="card"><div class="sec-lbl">INSIGHT SUMMARY</div><div style="font-size:17px;color:#e2ecf8;line-height:1.75">{r["summary"]}</div><div style="font-size:11px;color:#4a6080;margin-top:8px">All counts and quotes drawn directly from uploaded data ({TOTAL} valid respondents). No inference.</div></div>',unsafe_allow_html=True)
T=r["T"]
 
if r["is_comp"] and r["comp"]:
cp=r["comp"]; va,vb=cp["va"],cp["vb"]; na,nb=cp["na"],cp["nb"]
va_c=VENDOR_COLORS.get(va,"#f59e0b"); vb_c=VENDOR_COLORS.get(vb,"#93c5fd")
st.markdown(f'<div style="display:flex;gap:12px;margin-bottom:16px;align-items:center"><span style="background:{va_c}22;color:{va_c};border:1px solid {va_c}44;padding:2px 12px;border-radius:99px;font-size:12px;font-weight:600">{va} (n={na})</span><span style="color:#4a6080">vs</span><span style="background:{vb_c}22;color:{vb_c};border:1px solid {vb_c}44;padding:2px 12px;border-radius:99px;font-size:12px;font-weight:600">{vb} (n={nb})</span></div>',unsafe_allow_html=True)
st.markdown('<div class="card"><div class="sec-lbl">THEME COMPARISON</div>',unsafe_allow_html=True)
for row in cp["rows"]:
t=row["Theme"]; pa,pb,d=row["_pa"],row["_pb"],row["_d"]; color=TC.get(t,"#4a6080")
dh=f'<span class="diff-pos">▲{d}pp</span>' if d>0 else (f'<span class="diff-neg">▼{abs(d)}pp</span>' if d<0 else '<span style="color:#4a6080">═</span>')
st.markdown(f"""<div style="margin-bottom:12px">
<div style="display:flex;justify-content:space-between;font-size:12px;margin-bottom:3px"><span style="color:#e2ecf8;font-weight:500">{t}</span><span>{dh}</span></div>
<div style="display:flex;gap:6px;align-items:center;margin-bottom:2px"><span style="font-size:10px;color:{va_c};min-width:80px">{va}</span><div class="bar-trk" style="flex:1"><div class="bar-fill" style="width:{pa}%;background:{va_c}"></div></div><span style="font-family:'IBM Plex Mono',monospace;font-size:11px;color:{va_c};min-width:38px;text-align:right">{pa}%</span></div>
<div style="display:flex;gap:6px;align-items:center"><span style="font-size:10px;color:{vb_c};min-width:80px">{vb}</span><div class="bar-trk" style="flex:1"><div class="bar-fill" style="width:{pb}%;background:{vb_c}"></div></div><span style="font-family:'IBM Plex Mono',monospace;font-size:11px;color:{vb_c};min-width:38px;text-align:right">{pb}%</span></div>
</div>""",unsafe_allow_html=True)
st.markdown('</div>',unsafe_allow_html=True)
c1,c2=st.columns(2)
for col,df_s,val,vc3 in [(c1,cp["dfa"],va,va_c),(c2,cp["dfb"],vb,vb_c)]:
with col:
st.markdown(f'<div class="sec-lbl" style="color:{vc3}">{val} FULL RESPONSES</div>',unsafe_allow_html=True)
f=cp.get("focus",[])
filt=df_s[df_s["text_lower"].apply(lambda x: any(p in x for p in f) if f else True)]
for _,row in filt.head(4).iterrows(): quote_card(row)
else:
l,ri=st.columns([1,1])
with l:
cc1,cc2=st.columns(2)
cc1.markdown(f'<div class="card" style="text-align:center"><div class="stat-num">{r["n"]}</div><div class="stat-lbl">matched</div></div>',unsafe_allow_html=True)
cc2.markdown(f'<div class="card" style="text-align:center"><div class="stat-num">{round(r["n"]/T*100) if T else 0}%</div><div class="stat-lbl">of total</div></div>',unsafe_allow_html=True)
for title,data in r["chart"].items():
if data:
st.markdown(f'<div class="card"><div class="sec-lbl">{title}</div>',unsafe_allow_html=True)
for lbl,cnt in list(data.items())[:12]: bar_html(lbl,int(cnt),T,TC.get(lbl,"#3b6ef7"))
st.markdown('</div>',unsafe_allow_html=True)
with ri:
st.markdown('<div class="sec-lbl">FULL RESPONSES — ALL THEMES HIGHLIGHTED</div>',unsafe_allow_html=True)
st.markdown(f'<div style="font-size:11px;color:#4a6080;margin-bottom:12px">Showing {min(len(r["rows"]),6)} of {len(r["rows"])} matched respondents.</div>',unsafe_allow_html=True)
if r["rows"]:
for rec in r["rows"][:6]: quote_card(rec, r["topics"] if r["topics"] else None)
else:
st.markdown('<div style="color:#4a6080;font-size:13px">No matches. Try broader keywords.</div>',unsafe_allow_html=True)
if r.get("export"):
st.markdown("<hr>",unsafe_allow_html=True)
dfe=pd.DataFrame(r["export"])
st.dataframe(dfe[["ID","Practice_Setting","Specialty","CGP_Vendor"]].head(20),hide_index=True,use_container_width=True)
st.download_button("⬇ Download full responses CSV",dfe.to_csv(index=False).encode(),f"responses_{query[:20].replace(' ','_')}.csv","text/csv")
else:
st.markdown('<div style="text-align:center;margin-top:60px"><div style="font-size:48px;margin-bottom:16px">🧬</div><div style="font-size:16px;color:#4a6080">Upload your Excel file in the sidebar to begin</div><div style="font-size:13px;color:#2a3a55;margin-top:8px">Then use the sidebar tabs — Dashboard · Questions · Compare · Sentiment</div></div>',unsafe_allow_html=True)

